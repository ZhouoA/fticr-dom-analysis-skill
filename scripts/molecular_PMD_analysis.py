#!/usr/bin/env python
"""PMD network molecule matching and reaction-category statistics.

The workflow reads network_edge*.csv files, matches Source and Target formulas
against the corresponding FT-ICR DOM molecular-property workbook, and writes
row-preserving match tables plus VK/Group reaction statistics.
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font


VK_CATEGORIES = [
    "Lipids",
    "Aliphatic",
    "Lignin",
    "Carbohydrates",
    "Unsaturated",
    "Aromatic",
    "Tannin",
]
GROUP_CATEGORIES = ["CHO", "CHON", "CHONS", "CHOS"]
REACTION_GROUPS = {
    "1-CH": ["1-C3H6", "1-C3H4", "1-C2H4", "1-C2H2", "1-CH2", "1-C", "1-H2"],
    "1+CH": ["1+H2"],
    "1-CHO": ["1-CO", "1-CO2", "1-C2H2O2", "1-CH4O", "1-C6H8O6"],
    "1+CHO": ["1+C6H8O6", "1+C2H2O", "1+O", "1-H2+O", "1+H2O", "1+O2", "1-H2+O2"],
    "1-CHON": ["1-NH", "1-NH+O", "1-C2H5N", "1-CHON"],
    "1+CHON": ["1+NH", "1-O+NH", "1-O+NH3", "1-H+NO", "1-H+NO2"],
    "1-CHOS": ["1-SO2", "1-SO3"],
    "1+CHOS": ["1+SO3"],
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Build Source/Target PMD match tables and VK/Group reaction "
            "statistics from network_edge files."
        )
    )
    parser.add_argument(
        "input_dir",
        help="Directory containing network_edge*.csv, reaction_delta.csv, and analysis workbooks.",
    )
    parser.add_argument(
        "--output-dir",
        default=None,
        help="Output directory. Defaults to input_dir.",
    )
    parser.add_argument(
        "--edge-pattern",
        default="network_edge*.csv",
        help="Glob pattern for edge tables. Default: network_edge*.csv",
    )
    parser.add_argument(
        "--analysis-pattern",
        default="生物段{tag}_fticr_dom_analysis.xlsx",
        help="Analysis workbook pattern where {tag} comes from network_edge{tag}.csv.",
    )
    parser.add_argument(
        "--reaction-delta",
        default="reaction_delta.csv",
        help="Reaction-order table. It must contain a reaction column.",
    )
    parser.add_argument(
        "--include-all",
        action="store_true",
        help="Also include an ALL combined Dataset in statistics. Default: disabled.",
    )
    return parser.parse_args()


def normalize_text(series: pd.Series) -> pd.Series:
    return series.astype("string").str.strip()


def load_reaction_order(path: Path) -> list[str]:
    reaction_delta = pd.read_csv(path)
    if "reaction" not in reaction_delta.columns:
        raise ValueError(f"{path} must contain a reaction column.")
    reactions = normalize_text(reaction_delta["reaction"]).dropna().tolist()
    if not reactions:
        raise ValueError(f"{path} does not contain any reaction values.")
    return reactions


def load_analysis(path: Path) -> pd.DataFrame:
    frames: list[pd.DataFrame] = []
    with pd.ExcelFile(path) as xls:
        for sheet_name in xls.sheet_names:
            frame = pd.read_excel(xls, sheet_name=sheet_name)
            if "MolForm" not in frame.columns:
                raise ValueError(f"{path.name} sheet {sheet_name!r} does not contain MolForm.")
            frame = frame.copy()
            frame.insert(1, "AnalysisSheet", sheet_name)
            frame["MolForm"] = normalize_text(frame["MolForm"])
            frames.append(frame)

    analysis = pd.concat(frames, ignore_index=True)
    duplicated = analysis["MolForm"].duplicated(keep=False)
    if duplicated.any():
        examples = analysis.loc[duplicated, "MolForm"].drop_duplicates().head(10).tolist()
        raise ValueError(f"{path.name} contains duplicated MolForm values, examples: {examples}")
    return analysis


def tag_from_edge_path(edge_path: Path) -> str:
    return edge_path.stem.replace("network_edge", "", 1)


def build_side_match(
    edge_path: Path,
    analysis: pd.DataFrame,
    side: str,
    dataset: str,
) -> pd.DataFrame:
    edge = pd.read_csv(edge_path)
    for column in [side, "Reaction"]:
        if column not in edge.columns:
            raise ValueError(f"{edge_path.name} must contain {column}.")

    side_reaction = edge[[side, "Reaction"]].copy()
    side_reaction.insert(0, "NetworkEdgeOrder", range(1, len(side_reaction) + 1))
    side_reaction.insert(1, "Dataset", dataset)
    side_reaction[side] = normalize_text(side_reaction[side])
    side_reaction["Reaction"] = normalize_text(side_reaction["Reaction"])

    renamed_analysis = analysis.rename(columns={"MolForm": side})
    matched = side_reaction.merge(renamed_analysis, on=side, how="left", validate="many_to_one")
    matched.insert(4, "MatchedInAnalysis", matched["AnalysisSheet"].notna())
    return matched


def summarize_category(
    frame: pd.DataFrame,
    reaction_order: list[str],
    category_column: str,
    categories: list[str],
    dataset: str,
) -> pd.DataFrame:
    working = frame.copy()
    working["RI"] = pd.to_numeric(working.get("RI"), errors="coerce").fillna(0.0)
    rows: list[dict[str, object]] = []

    for reaction in reaction_order:
        subset = working[working["Reaction"] == reaction]
        row: dict[str, object] = {"Dataset": dataset, "Reaction": reaction}
        for category in categories:
            category_subset = subset[subset[category_column] == category]
            row[f"{category}_count"] = int(len(category_subset))
            row[f"{category}_RI_sum（%）"] = float(category_subset["RI"].sum() * 100)
        rows.append(row)
    return pd.DataFrame(rows)


def append_reaction_group_and_dataset_sums(frame: pd.DataFrame) -> pd.DataFrame:
    numeric_columns = [
        column
        for column in frame.columns
        if str(column).endswith("_count") or str(column).endswith("RI_sum（%）")
    ]
    output_parts: list[pd.DataFrame] = []

    for dataset in frame["Dataset"].drop_duplicates().tolist():
        subset = frame[frame["Dataset"] == dataset].copy()
        output_parts.append(subset)

        for group, reactions in REACTION_GROUPS.items():
            group_subset = subset[subset["Reaction"].isin(reactions)]
            sum_row = {column: None for column in frame.columns}
            sum_row["Dataset"] = dataset
            sum_row["Reaction"] = f"{group}_sum"
            for column in numeric_columns:
                sum_row[column] = pd.to_numeric(group_subset[column], errors="coerce").sum()
            output_parts.append(pd.DataFrame([sum_row], columns=frame.columns))

        dataset_sum = {column: None for column in frame.columns}
        dataset_sum["Dataset"] = f"{dataset}_sum"
        dataset_sum["Reaction"] = f"{dataset}_sum"
        for column in numeric_columns:
            dataset_sum[column] = pd.to_numeric(subset[column], errors="coerce").sum()
        output_parts.append(pd.DataFrame([dataset_sum], columns=frame.columns))

    return pd.concat(output_parts, ignore_index=True)


def write_statistics(
    path: Path,
    matched_frames: list[tuple[str, pd.DataFrame]],
    reaction_order: list[str],
    include_all: bool,
) -> None:
    vk_stats: list[pd.DataFrame] = []
    group_stats: list[pd.DataFrame] = []

    for dataset, matched in matched_frames:
        vk_stats.append(summarize_category(matched, reaction_order, "VK", VK_CATEGORIES, dataset))
        group_stats.append(summarize_category(matched, reaction_order, "Group", GROUP_CATEGORIES, dataset))

    if include_all and matched_frames:
        all_matched = pd.concat([matched for _, matched in matched_frames], ignore_index=True)
        vk_stats.append(summarize_category(all_matched, reaction_order, "VK", VK_CATEGORIES, "ALL"))
        group_stats.append(summarize_category(all_matched, reaction_order, "Group", GROUP_CATEGORIES, "ALL"))

    vk_output = append_reaction_group_and_dataset_sums(pd.concat(vk_stats, ignore_index=True))
    group_output = append_reaction_group_and_dataset_sums(pd.concat(group_stats, ignore_index=True))

    with pd.ExcelWriter(path) as writer:
        vk_output.to_excel(writer, index=False, sheet_name="VK_stats")
        group_output.to_excel(writer, index=False, sheet_name="Group_stats")
    bold_sum_rows(path)


def bold_sum_rows(path: Path) -> None:
    workbook = load_workbook(path)
    bold_font = Font(bold=True)

    for worksheet in workbook.worksheets:
        headers = [cell.value for cell in worksheet[1]]
        dataset_col = headers.index("Dataset") + 1
        reaction_col = headers.index("Reaction") + 1
        for row_number in range(2, worksheet.max_row + 1):
            dataset_value = worksheet.cell(row_number, dataset_col).value
            reaction_value = worksheet.cell(row_number, reaction_col).value
            if (
                isinstance(dataset_value, str)
                and dataset_value.endswith("_sum")
            ) or (
                isinstance(reaction_value, str)
                and reaction_value.endswith("_sum")
            ):
                for cell in worksheet[row_number]:
                    cell.font = bold_font

    workbook.save(path)


def run_analysis(args: argparse.Namespace) -> list[dict[str, object]]:
    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir) if args.output_dir else input_dir
    source_dir = output_dir / "source_reaction_matches"
    target_dir = output_dir / "target_reaction_matches"
    source_dir.mkdir(parents=True, exist_ok=True)
    target_dir.mkdir(parents=True, exist_ok=True)

    reaction_order = load_reaction_order(input_dir / args.reaction_delta)
    source_frames: list[tuple[str, pd.DataFrame]] = []
    target_frames: list[tuple[str, pd.DataFrame]] = []
    summary: list[dict[str, object]] = []

    edge_paths = sorted(input_dir.glob(args.edge_pattern))
    if not edge_paths:
        raise FileNotFoundError(f"No edge tables matched {args.edge_pattern!r} in {input_dir}.")

    for edge_path in edge_paths:
        dataset = tag_from_edge_path(edge_path)
        analysis_path = input_dir / args.analysis_pattern.format(tag=dataset)
        if not analysis_path.exists():
            raise FileNotFoundError(f"Missing analysis workbook for {dataset}: {analysis_path}")

        analysis = load_analysis(analysis_path)
        source_match = build_side_match(edge_path, analysis, "Source", dataset)
        target_match = build_side_match(edge_path, analysis, "Target", dataset)
        source_frames.append((dataset, source_match))
        target_frames.append((dataset, target_match))

        source_output = source_dir / f"source_reaction_{dataset}_matched_analysis.xlsx"
        target_output = target_dir / f"target_reaction_{dataset}_matched_analysis.xlsx"
        source_match.to_excel(source_output, index=False)
        target_match.to_excel(target_output, index=False)

        summary.append(
            {
                "dataset": dataset,
                "edge_rows": int(len(pd.read_csv(edge_path))),
                "source_matched_rows": int(source_match["MatchedInAnalysis"].sum()),
                "source_unmatched_rows": int((~source_match["MatchedInAnalysis"]).sum()),
                "target_matched_rows": int(target_match["MatchedInAnalysis"].sum()),
                "target_unmatched_rows": int((~target_match["MatchedInAnalysis"]).sum()),
                "source_output": str(source_output),
                "target_output": str(target_output),
            }
        )

    source_stats = source_dir / "source_reaction_VK_Group_statistics.xlsx"
    target_stats = target_dir / "target_reaction_VK_Group_statistics.xlsx"
    write_statistics(source_stats, source_frames, reaction_order, args.include_all)
    write_statistics(target_stats, target_frames, reaction_order, args.include_all)

    summary.append({"source_statistics": str(source_stats), "target_statistics": str(target_stats)})
    summary_path = output_dir / "molecular_PMD_analysis_summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    return summary


def main() -> int:
    args = parse_args()
    try:
        summary = run_analysis(args)
    except Exception as exc:
        print(f"Error: {exc}")
        return 1
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
