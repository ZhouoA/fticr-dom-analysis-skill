from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_ZERO_COLUMNS = ("Br",)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert three FT-ICR MS Excel workbooks to UTF-8 CSV for U-S-V图."
    )
    parser.add_argument("--input_dir", required=True)
    parser.add_argument("--output_dir", required=True)
    parser.add_argument("--sample_ids", default="YL,ML,OL")
    return parser.parse_args()


def convert_workbook(
    sample_id: str,
    input_dir: Path,
    output_dir: Path,
) -> None:
    input_path = input_dir / f"{sample_id}.xlsx"
    output_path = output_dir / f"{sample_id}.csv"
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = worksheet.iter_rows(values_only=True)
    header = [str(value) if value is not None else "" for value in next(rows)]

    for column in REQUIRED_ZERO_COLUMNS:
        if column not in header:
            header.append(column)

    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=header)
        writer.writeheader()
        original_columns = header[: worksheet.max_column]
        for values in rows:
            record = dict(zip(original_columns, values))
            for column in REQUIRED_ZERO_COLUMNS:
                record.setdefault(column, 0)
            writer.writerow(record)

    workbook.close()
    print(f"{sample_id}: {worksheet.max_row - 1} rows -> {output_path}")


def main() -> None:
    args = parse_args()
    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)
    sample_ids = tuple(
        sample_id.strip()
        for sample_id in args.sample_ids.split(",")
        if sample_id.strip()
    )
    if len(sample_ids) != 3:
        raise ValueError("--sample_ids must contain exactly three comma-separated IDs.")
    output_dir.mkdir(parents=True, exist_ok=True)
    for sample_id in sample_ids:
        convert_workbook(sample_id, input_dir, output_dir)


if __name__ == "__main__":
    main()
